import pandas as pd
from datetime import date
#import xlsxwriter
from openpyxl import load_workbook
import time
from win32com import client
import pywintypes
import math

#from fpdf import FPDF
#pdf = FPDF()

# Collect THD Order Details
#ord_data = pd.read_excel("Y:/User/Order Entry User/CTM Household Appliances Orders Processed/The Home Depot Inc/FILES/Daily Order/THD_Order_DTLS_" + str(date.today()) + ".xlsx", sheet_name='Sheet1')
ord_data = pd.read_excel("Y:/User/Order Entry User/CTM Household Appliances Orders Processed/The Home Depot Inc/FILES/Daily Order/THD_Order_DTLS_2022-11-04.xlsx", sheet_name='Sheet1')
#NetSuite Item Child Relationships
ns_item_dir = 'Y:/User/Order Entry User/NETSUITE IMPORTS/NETSUITE ITEMS'
itemdf = pd.read_excel(ns_item_dir + "/Item_Child_Rel.xlsx")
ord_data = ord_data.merge(itemdf, on="Item ID", how="left")

ns_loc_dir = 'Y:/User/Order Entry User/NETSUITE IMPORTS/LOCATION'
locdf = pd.read_csv(ns_loc_dir + "/NS_Location.csv")
ord_data = ord_data.merge(locdf, on="Location ID", how="left")
ord_data = ord_data.sort_values(by='External ID', ascending=False)
ord_data = ord_data.reset_index()
ord_data = ord_data.drop(columns=['index'])  # Drop old index
ord_data['C Item ID'] = ord_data['C Item ID'].astype(str)
ord_data['PO Number'] = ord_data['PO Number'].astype(str)
ord_data['PO Number'] = ord_data['PO Number'].str.strip()  # PO# # IF NOT 8 DIGITS ADD 0 IN FRONT
ord_data['PO Number'] = ord_data['PO Number'].str.zfill(8)  # PO# # IF NOT 8 DIGITS ADD 0 IN FRONT
ord_data['Zip'] = ord_data['Zip'].astype(str)
ord_data['Zip'] = ord_data['Zip'].str.strip()  # PO# # IF NOT 8 DIGITS ADD 0 IN FRONT
ord_data['Zip'] = ord_data['Zip'].str.zfill(5)  # PO# # IF NOT 8 DIGITS ADD 0 IN FRONT
ord_data['ShipTo Postal Code'] = ord_data['ShipTo Postal Code'].astype(str)
ord_data['ShipTo Postal Code'] = ord_data['ShipTo Postal Code'].str.strip()  # PO# # IF NOT 8 DIGITS ADD 0 IN FRONT
ord_data['ShipTo Postal Code'] = ord_data['ShipTo Postal Code'].str.zfill(5)  # PO# # IF NOT 8 DIGITS ADD 0 IN FRONT
print("Hello")

files_written = pd.DataFrame(columns=['FileName'], index=range(len(ord_data)))

for i in range(len(ord_data)):

    if i == 0 or (ord_data['External ID'].iloc[i] != ord_data['External ID'].iloc[(i-1)]): #FIRST ENTRY UES THE SHIP FROM LOCATION AND SHIP TO LOCATION NO MATTER WHAT
        file_path = "Y:/User/Order Entry User/CTM Household Appliances Orders Processed/The Home Depot Inc/EMPTY BOL.xlsm"
        #wb = load_workbook(file_path)
        wb = load_workbook(filename=file_path, read_only=False, keep_vba=True)
        ws = wb['Table 1']  # or wb.active
        # SHIP FROM
        ws['C3'] = "CTM Household Appliances (" + ord_data['City'].iloc[i] +")" # SHIP FROM LOCATION NAME
        ws['C5'] = ord_data['Address 1'].iloc[i]  # SHIP FORM ADDRESS
        ws['C6'] = (ord_data['City'].iloc[i]+", "+ord_data['State/Province'].iloc[i]+", "+ord_data['Zip'].iloc[i])  # SHIP FORM ADDRESS
        ws['C7'] = ord_data['Phone'].iloc[i]  # SHIP FROM PHONE
        # CUSTOMER PO
        ws['J3'] = ord_data['PO Number'].iloc[i]  # SHIP PO#
        # SHIP TO
        ws['C9'] = ord_data['ShipTo Name'].iloc[i]  # SHIP TO NAME
        ws['C10'] = ord_data['ShipTo Address1'].iloc[i]  # SHIP TO ADDRESS
        if ord_data['ShipTo Address2'].iloc[i] != "N/A":
            ws['C11'] = ord_data['ShipTo Address2'].iloc[i]  # SHIP TO ADDRESS2
        ws['C12'] = (ord_data['ShipTo City'].iloc[i] + ", " + ord_data['ShipTo State_x'].iloc[i] + ", " + ord_data['ShipTo Postal Code'].iloc[i])  # SHIP TO ADDRESS
        ws['C13'] = str(ord_data['ShipTo Day Phone'].iloc[i])
        if ord_data['ShipTo Address Rate Class'].iloc[i] != "RESIDENTIAL":
            ws['I18'] = "Ref Name: " + ord_data['BillTo Name'].iloc[i]
        else:
            ws['C17'] = "LIFTGATE APPROVED"
        # CARRIER NAME
        ws['I8'] = ord_data['CARRIER NAME'].iloc[i]
        # ITEM
        ws['C22'] = ord_data['Quantity'].iloc[i]
        if ord_data['C Item ID'].iloc[i] == "nan":  # IF THERE IS NO CHILD ITEM
            ws['E22'] = ord_data['Vendor SKU'].iloc[i]
            ws['D22'] = ord_data['Item Weight'].iloc[i]
            ws['K22'] = ord_data['BOL Class'].iloc[i]
        else:  # IF THERE IS A CHILD ITEM
            ws['E22'] = ord_data['C Item ID'].iloc[i]
            ws['D22'] = ord_data['C Item Weight'].iloc[i]
            ws['H22'] = "Part of SKU#" + str(ord_data['Vendor SKU'].iloc[i])
            ws['K22'] = ord_data['BOL Class'].iloc[i]
        files_written.iloc[i, 0] = "Y:/User/Order Entry User/CTM Household Appliances Orders Processed/The Home Depot Inc/NEW - PROCESS/THD BOL PO#"+ord_data['PO Number'].iloc[i] + ord_data['Location'].iloc[i] + ".xlsm"
        wb.save(files_written.iloc[i, 0])
    elif ord_data['External ID'].iloc[i] == ord_data['External ID'].iloc[(i-1)]:
        file_path = "Y:/User/Order Entry User/CTM Household Appliances Orders Processed/The Home Depot Inc/NEW - PROCESS/THD BOL PO#"+ord_data['PO Number'].iloc[i] + ord_data['Location'].iloc[i] + ".xlsm"
        #wb = load_workbook(file_path)
        wb = load_workbook(filename=file_path, read_only=False, keep_vba=True)
        ws = wb['Table 1']  # or wb.active
        idx = (ord_data['External ID'] == ord_data['External ID'].iloc[i]).idxmax() # INDEX of First Occurrence of the PO#
        itemno = i-idx  #item number per order
        lineno = 22+itemno #excel line number to write on
        ws['C'+str(lineno)] = ord_data['Quantity'].iloc[i]
        if ord_data['C Item ID'].iloc[i] == "nan":  # IF THERE IS NO CHILD ITEM
            ws['E'+str(lineno)] = ord_data['Vendor SKU'].iloc[i]
            ws['D'+str(lineno)] = ord_data['Item Weight'].iloc[i]
            ws['K' + str(lineno)] = ord_data['BOL Class'].iloc[i]
        else:  # IF THERE IS A CHILD ITEM
            ws['E'+str(lineno)] = ord_data['C Item ID'].iloc[i]
            ws['H' + str(lineno)] = "Part of SKU#" + str(ord_data['Vendor SKU'].iloc[i])
            ws['D'+str(lineno)] = ord_data['C Item Weight'].iloc[i]
            ws['K' + str(lineno)] = ord_data['BOL Class'].iloc[i]
        files_written.iloc[i, 0] = "Y:/User/Order Entry User/CTM Household Appliances Orders Processed/The Home Depot Inc/NEW - PROCESS/THD BOL PO#"+ord_data['PO Number'].iloc[i] + ord_data['Location'].iloc[i] + ".xlsm"
        wb.save(files_written.iloc[i, 0])
time.sleep(1)
print("done")

import os, os.path
import win32com.client as wincl

files_written = files_written.drop_duplicates()
for i in range(len(files_written)):
    excel_macro = wincl.DispatchEx("Excel.application")
    excel_path = os.path.expanduser(files_written.iloc[i, 0])
    workbook = excel_macro.Workbooks.Open(Filename=excel_path, ReadOnly=1)
    excel_macro.Application.Run("Save_PDF_Current_Folder")
    excel_macro.Application.Quit()

#ord_data.to_excel(r"Y:/User/Order Entry User/CTM Household Appliances Orders Processed/The Home Depot Inc/THD_"
#        + str(date.today()) + ".xlsx", index=False, header=True)

print("hello")

for i in range(len(ord_data)):
    if ord_data['C Item ID'][i] == "nan":
        ord_data['C Item ID'][i] = ord_data['Vendor SKU'][i]


ord_data['C Item ID'][0] = ""
ord_data['coalesce'] = ord_data[['C Item ID', 'Vendor SKU']].bfill(axis=1).iloc[:, 0]